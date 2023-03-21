import openpyxl
import pandas as pd
import numpy as np

from openpyxl import load_workbook
wb = load_workbook('input.xlsx')
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
    for value in row:    
        print(value,"\t",end='')
    print("")